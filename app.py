from openpyxl import Workbook, load_workbook
from flask import Flask, render_template, request, redirect, url_for
from datetime import datetime

app = Flask(__name__)

...

# Pegando data atual
dia = datetime.now().day
mes = datetime.now().month
ano = datetime.now().year

if mes == 1:
    mesEscrito = 'Janeiro'
elif mes == 2:
    mesEscrito = 'Fevereiro'
elif mes == 3:
    mesEscrito = 'Março'
elif mes == 4:
    mesEscrito = 'Abril'
elif mes == 5:
    mesEscrito = 'Maio'
elif mes == 6:
    mesEscrito = 'Junho'
elif mes == 7:
    mesEscrito = 'Julho'
elif mes == 8:
    mesEscrito = 'Agosto'
elif mes == 9:
    mesEscrito = 'Setembro'
elif mes == 10:
    mesEscrito = 'Outubro'
elif mes == 11:
    mesEscrito = 'Novembro'
elif mes == 12:
    mesEscrito = 'Dezembro'

...

# Carregando planilha
try:
    wb = load_workbook(f'./planilhas/{mesEscrito} {ano}.xlsx', data_only=True)
except FileNotFoundError:
    wb = load_workbook(f'Base.xlsx', data_only=True)
    # wb = Workbook()
    # ws = wb.active
    # ws.title = f'Dia {dia}'
    # ws.append(["Dinheiro", "Débito", "Crédito", "Parcelas"])

...

# Criando aba para soma
# try:
#     ws = wb['Somas']
# except:
#     ws = wb.create_sheet('Somas')
#     ws.append([" ", "Dinheiro", "Débito", "Crédito", " ", "Total"])
#     total = { 'Dinheiro': 0, 'Debito': 0, 'Credito': 0, 'Total': 0 }
#     ws.append([f"Dia {dia}", total['Dinheiro'], total['Debito'], total['Credito'], " ", total['Total']])

# ...

try:
    ws = wb[f'Soma']
    total = { 'Dinheiro': 0, 'Debito': 0, 'Credito': 0, 'Total': 0 }
    total = { 'Dinheiro': ws[f'B{dia+1}'], 'Debito': ws[f'C{dia+1}'], 'Credito': ws[f'D{dia+1}'], 'Total': ws[f'F{dia+1}'] }
except:
    print('Erro')

# Acessando aba
try:
    ws = wb[f'Dia {dia}']
except:
    ws = wb.create_sheet(f'Dia {dia}')
    ws.append(["Dinheiro", "Débito", "Crédito", "Parcelas"])

...

# Função para salvar o arquivo
def salvar():
    try:
        wb.save(f'./planilhas/{mesEscrito} {ano}.xlsx')
    except PermissionError:
        print('Feche a planilha para salvar!')

...

# Função para registrar venda no Dinheiro ou Débito
def venda_D(valor, metodo):
    valor = int(valor)
    for c in range(1, 200):
        if metodo == 'Dinheiro':
            celula = ws[f'A{c}'].value
            if celula is None or celula == '':
                ws[f'A{c}'] = valor
                total['Dinheiro'].value += valor
                total['Total'].value += valor
                print(total['Dinheiro'])
                break
        elif metodo == 'Debito':
            celula = ws[f'B{c}'].value
            if celula is None or celula == '':
                ws[f'B{c}'] = valor
                total['Debito'].value += valor
                total['Total'].value += valor
                break

    # Salvando arquivo
    salvar()

...

# Função para registrar venda no Crédito
def venda_C(valor, parcelas):
    valor = int(valor)
    for c in range(1, 200):
        celula = ws[f'C{c}'].value
        if celula is None or celula == '':
            ws[f'C{c}'] = valor
            ws[f'D{c}'] = parcelas
            total['Credito'].value += valor
            total['Total'].value += valor
            break
    
    # Salvando arquivo
    salvar()

...

# Página inicial: Registrar vendas
@app.route('/')
def index():
    return render_template('index.html')

...

# Página para visualizar a planilha do dia
@app.route('/planilha_dia')
def planilha_dia():
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)

    return render_template('planilha_dia.html', data=data)

...

# Página para visualizar a planilha do mês
@app.route('/planilha_mes')
def planilha_mes():
    data = []
    ws = wb['Soma']
    for row in ws.iter_rows(values_only=True):
        data.append(row)

    return render_template('planilha_mes.html', data=data)

...

# Registrando venda
@app.route('/registro', methods=['POST'])
def registro():
    valor = request.form['valor']
    metodo = request.form['metodo']

    if metodo == 'Credito':
        parcelas = request.form['parcelas']
        venda_C(valor, parcelas)
    else:
        venda_D(valor, metodo)

    return redirect(url_for('index'))

...

# Rodando o programa
if __name__ == '__main__':
    app.run(debug=True, port=5000)
