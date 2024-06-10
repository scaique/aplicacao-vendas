from flask import flash
from openpyxl import load_workbook
from .date import obter_data_atual
import os

mesEscrito = obter_data_atual()[3]
ano = obter_data_atual()[2]

def carregar_planilha():
    try:
        wb = load_workbook(f'./planilhas/{ano}/{mesEscrito} {ano}.xlsx', data_only=True)
    except FileNotFoundError:
        wb = load_workbook(f'./Base.xlsx', data_only=True)
    return wb

def calculo():
    dia = obter_data_atual()[0]

    try:
        wb = carregar_planilha()
    except Exception as e:
        return e

    try:
        ws = wb[f'Dia {dia}']
    except Exception as e:
        ws = wb.create_sheet(f'Dia {dia}')
        ws.append(["Dinheiro", "Débito", "Crédito", "Parcelas"])

    try:
        soma = wb[f'Dia {dia}']
        somaDin = 0
        somaDeb = 0
        somaCre = 0

        for cell in soma['A'][1:]:
            if cell.value is not None:
                somaDin += cell.value
        for cell in soma['B'][1:]:
            if cell.value is not None:
                somaDeb += cell.value
        for cell in soma['C'][1:]:
            if cell.value is not None:
                somaCre += cell.value

        total = {
            'Dinheiro': somaDin,
            'Debito': somaDeb,
            'Credito': somaCre,
            'Total': somaDin+somaDeb+somaCre
        }
    except Exception as e:
        return e
    
    return total

def wb_ws_total():
    dia = obter_data_atual()[0]

    try:
        wb = carregar_planilha()
        ws = wb[f'Soma']
        total = calculo()
        
        ws[f'B{dia+1}'] = total['Dinheiro']
        ws[f'C{dia+1}'] = total['Debito']
        ws[f'D{dia+1}'] = total['Credito']
        ws[f'F{dia+1}'] = total['Total']
    except Exception as e:
        return e

    try:
        ws = wb[f'Dia {dia}']
    except Exception as e:
        ws = wb.create_sheet(f'Dia {dia}')
        ws.append(["Dinheiro", "Débito", "Crédito", "Parcelas"])
    
    return wb, ws, total

def salvar(wb):
    if not os.path.exists(f'./planilhas'):
        flash("Criando diretório 'planilhas'...", 'success')
        os.mkdir(f'./planilhas')
    if not os.path.exists(f'./planilhas/{ano}'):
        flash(f"Criando diretório 'planilhas/{ano}'...", 'success')
        os.mkdir(f'./planilhas/{ano}')
    try:
        wb.save(f'./planilhas/{ano}/{mesEscrito} {ano}.xlsx')
    except PermissionError:
        flash('Erro ao salvar, talvez você precise fechar a planilha!', 'success')

def venda_D(ws, wb, total, valor, metodo):
    valores = valor.strip().split()
    for val in valores:
        val = int(val)
        flash(f'R${val},00 no {metodo} registrado com sucesso.', 'success')
        for c in range(1, 500):
            if metodo == 'Dinheiro':
                celula = ws[f'A{c}'].value
                if celula is None or celula == '':
                    ws[f'A{c}'] = val
                    total['Dinheiro'] += val
                    total['Total'] += val
                    break
            elif metodo == 'Debito':
                celula = ws[f'B{c}'].value
                if celula is None or celula == '':
                    ws[f'B{c}'] = val
                    total['Debito'] += val
                    total['Total'] += val
                    break
    salvar(wb)

def venda_C(ws, wb, total, valor, parcelas):
    valores = valor.strip().split()
    for val in valores:
        val = int(val)
        flash(f'R${val},00 em {parcelas} no Crédito registrado com sucesso.', 'success')
        for c in range(1, 500):
            celula = ws[f'C{c}'].value
            if celula is None or celula == '':
                ws[f'C{c}'] = val
                ws[f'D{c}'] = parcelas
                total['Credito'] += val
                total['Total'] += val
                break
    salvar(wb)

def troco_dia(ws, wb, valor, dia):
    valor = int(valor)
    ws = wb['Soma']
    ws[f'H{dia+1}'] = valor
    ws[f'H{dia+2}'] = valor
    if ws['H33'].value != 'TOTAL:':
        ws[f'H33'].value = 'TOTAL:'
    salvar(wb)
    flash(f'Troco no valor de R${valor},00 registrado com sucesso.', 'success')

def troco_mes(ws, wb, valor):
    valor = int(valor)
    ws = wb['Soma']
    ws['K2'] = valor
    ws['H2'] = valor
    salvar(wb)
    flash(f'Troco no valor de R${valor},00 registrado com sucesso.', 'success')

def calculo_total(ws, wb):
    ws = wb['Soma']
    dia = obter_data_atual()[0]
    total = wb_ws_total()[2]

    ws[f'B{dia+1}'] = total['Dinheiro']
    ws[f'C{dia+1}'] = total['Debito']
    ws[f'D{dia+1}'] = total['Credito']
    ws[f'F{dia+1}'] = total['Total']

    if ws['B2'].value is not None and ws['H2'].value is not None and ws['K2'].value is not None:
        total_dia1 = ws['B2'].value + ws['H2'].value - ws['K2'].value
        ws['I2'] = total_dia1

    for c in range(3, 33):
        if ws[f'B{c}'].value is not None and ws[f'H{c-1}'].value is not None and ws[f'H{c}'].value is not None:
            retirada = int(ws[f'B{c}'].value) + int(ws[f'H{c-1}'].value) - int(ws[f'H{c}'].value)
            ws[f'I{c}'] = retirada
    
    total_mes = 0
    
    for c in range(2, 33):
        if ws[f'I{c}'].value is not None:
            total_mes += int(ws[f'I{c}'].value)
            ws['H33'].value = 'TOTAL:'
            ws['I33'].value = total_mes

    salvar(wb)
